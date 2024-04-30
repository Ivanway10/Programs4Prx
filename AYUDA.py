from openpyxl import load_workbook
from openpyxl import Workbook

def main():
    # Cargar los archivos xlsx
    B_wb = load_workbook('opm.xlsx')
    A_wb = load_workbook('praxis.xlsx')

    # Seleccionar la hoja de trabajo en cada archivo
    B_sheet = B_wb.active
    A_sheet = A_wb.active

    # Crear un diccionario para almacenar los elementos de B
    B_elements = {}

    # Leer los elementos de B y guardarlos en el diccionario
    for row in B_sheet.iter_rows(min_row=2, values_only=True):
        element_id, last_name, first_name = row[0], row[1], row[2]
        B_elements[(last_name, first_name)] = element_id

    # Crear un archivo de resultados
    results_wb = Workbook()
    results_sheet = results_wb.active

    # Escribir encabezados en el archivo de resultados
    results_sheet.append(['First Name', 'Last Name', 'A ID', 'B ID'])

    # Leer los elementos de A y comparar con los de B
    for row in A_sheet.iter_rows(min_row=2, values_only=True):
        A_id, last_name, first_name = row[0], row[1], row[2]
        B_id = B_elements.get((last_name, first_name))

        if B_id and B_id != A_id:
            results_sheet.append([first_name, last_name, A_id, B_id])

    # Guardar el archivo de resultados
    results_wb.save('elementos_con_ids_distintos.xlsx')

if __name__ == "__main__":
    main()
